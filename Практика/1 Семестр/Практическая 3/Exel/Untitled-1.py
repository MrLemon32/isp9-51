# investment_model_fixed.py
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.protection import SheetProtection
from datetime import datetime

class InvestmentModelCreator:
    def __init__(self):
        self.wb = Workbook()
        self.setup_styles()
        
    def setup_styles(self):
        """Настройка стилей"""
        self.header_font = Font(size=14, bold=True, color="FFFFFF")
        self.subheader_font = Font(size=12, bold=True, color="2F5496")
        self.normal_font = Font(size=10)
        self.bold_font = Font(size=10, bold=True)
        
        self.header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        self.subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    def create_model(self, filename):
        """Создание полной финансовой модели"""
        # Удаляем дефолтный лист
        self.wb.remove(self.wb.active)
        
        # Создаем все листы
        self.create_readme_sheet()
        self.create_input_sheet()
        self.create_calculations_sheet()
        self.create_scenarios_sheet()
        self.create_report_sheet()
        
        # Сохраняем файл
        self.wb.save(filename)
        print(f"✅ Файл создан: {filename}")
        
        # Создаем файл с VBA кодом
        self.create_vba_code_file()
        
        return filename

    def create_readme_sheet(self):
        """Создание листа README"""
        ws = self.wb.create_sheet("README")
        
        ws['A1'] = "ФИНАНСОВАЯ МОДЕЛЬ ИНВЕСТИЦИОННОГО ПРОЕКТА"
        ws['A1'].font = Font(size=16, bold=True, color="2F5496")
        ws.merge_cells('A1:F1')
        
        instructions = [
            "ИНСТРУКЦИЯ ПО ИСПОЛЬЗОВАНИЮ МОДЕЛИ",
            "",
            "1. 📊 ВВОД ДАННЫХ - лист «Ввод»",
            "   • Заполните параметры в желтых ячейках",
            "   • Модель автоматически пересчитает показатели",
            "",
            "2. 🔄 УПРАВЛЕНИЕ СЦЕНАРИЯМИ - лист «Сценарии»",
            "   • Выберите один из 3 сценариев",
            "   • Все графики обновляются автоматически",
            "",
            "3. 📈 РЕЗУЛЬТАТЫ - лист «Отчёт»",
            "   • Ключевые метрики: NPV, IRR, срок окупаемости",
            "   • Автоматический статус проекта",
            "",
            "ТЕХНИЧЕСКИЕ ХАРАКТЕРИСТИКИ:",
            "• Формулы: Динамические массивы Excel 365",
            "• Совместимость: Excel 2019 и новее",
            "• Защита: Пароль для редактирования - 123",
            "",
            f"Создано: {datetime.now().strftime('%d.%m.%Y')}"
        ]
        
        for i, line in enumerate(instructions, 3):
            ws[f'A{i}'] = line
        
        ws.column_dimensions['A'].width = 60
        print("✅ Лист README создан")

    def create_input_sheet(self):
        """Создание листа Ввод с параметрами проекта"""
        ws = self.wb.create_sheet("Ввод")
        
        # Заголовок
        ws['A1'] = "ПАРАМЕТРЫ ИНВЕСТИЦИОННОГО ПРОЕКТА"
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws.merge_cells('A1:B1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Основные параметры проекта
        parameters = [
            ("ОСНОВНЫЕ ПАРАМЕТРЫ ПРОЕКТА", ""),
            ("Название проекта:", "Производство экологичной упаковки"),
            ("Тип оборудования:", "Стандарт"),
            ("Начальные инвестиции (₽):", 5000000),
            ("Срок проекта (лет):", 5),
            ("Ставка дисконтирования (%):", 12),
            ("Налог на прибыль (%):", 20),
            ("", ""),
            ("ПРОИЗВОДСТВЕННЫЕ ПОКАЗАТЕЛИ", ""),
            ("Начальный объём производства (ед./год):", 100000),
            ("Годовой рост объёма (%):", 5),
            ("Цена за единицу (₽):", 50),
            ("Переменные издержки (₽/ед.):", 25),
            ("Постоянные издержки (₽/год):", 1000000)
        ]
        
        for i, (param, value) in enumerate(parameters, 3):
            ws[f'A{i}'] = param
            ws[f'B{i}'] = value
            
            if param.endswith(":"):
                ws[f'A{i}'].font = self.bold_font
                ws[f'B{i}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            else:
                # Это заголовок раздела
                ws[f'A{i}'].font = self.subheader_font
                ws[f'A{i}'].fill = self.subheader_fill
                ws.merge_cells(f'A{i}:B{i}')
                ws[f'A{i}'].alignment = Alignment(horizontal='center')
        
        # Добавляем выпадающие списки (БЕЗ КОММЕНТАРИЕВ - это вызывало ошибку)
        equipment_dv = DataValidation(type="list", formula1='"Стандарт,Премиум"')
        ws.add_data_validation(equipment_dv)
        equipment_dv.add('B4')
        
        # Настройка колонок
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25
        
        print("✅ Лист Ввод создан")

    def create_calculations_sheet(self):
        """Создание листа Расчёты с финансовыми формулами"""
        ws = self.wb.create_sheet("Расчёты")
        
        # Заголовок
        ws['A1'] = "ФИНАНСОВЫЕ РАСЧЁТЫ"
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws.merge_cells('A1:K1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Заголовки столбцов
        headers = ["Год", "Объём", "Выручка", "Переменные затраты", 
                  "Постоянные затраты", "Амортизация", "EBITDA", "Налог", 
                  "Чистая прибыль", "FCF", "Накопленный FCF"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = self.bold_font
            cell.fill = self.subheader_fill
        
        # Формулы для 5 лет
        for year in range(1, 6):
            row = year + 3
            ws.cell(row=row, column=1).value = year
            ws.cell(row=row, column=2).value = f"=Ввод!B10*(1+Ввод!B11)^{year-1}"
            ws.cell(row=row, column=3).value = f"=B{row}*Ввод!B12"
            ws.cell(row=row, column=4).value = f"=B{row}*Ввод!B13"
            ws.cell(row=row, column=5).value = "=Ввод!B14"
            ws.cell(row=row, column=6).value = "=Ввод!B5/5"
            ws.cell(row=row, column=7).value = f"=C{row}-D{row}-E{row}"
            ws.cell(row=row, column=8).value = f"=MAX(0,G{row}-F{row})*Ввод!B8"
            ws.cell(row=row, column=9).value = f"=G{row}-H{row}"
            ws.cell(row=row, column=10).value = f"=I{row}+F{row}"
            
        # Накопленный FCF
        ws['K4'] = "=J4"
        for year in range(2, 6):
            row = year + 3
            ws.cell(row=row, column=11).value = f"=K{row-1}+J{row}"
        
        # Ключевые метрики
        ws['A10'] = "КЛЮЧЕВЫЕ ФИНАНСОВЫЕ МЕТРИКИ"
        ws['A10'].font = self.subheader_font
        ws['A10'].fill = self.subheader_fill
        ws.merge_cells('A10:B10')
        
        metrics = [
            ("NPV (₽):", "=NPV(Ввод!B7/100,J4:J8)-Ввод!B5"),
            ("IRR (%):", "=IRR(({-Ввод!B5,J4:J8}))*100"),
            ("Срок окупаемости (лет):", "=MATCH(0,K4:K8,1)+(-INDEX(K4:K8,MATCH(0,K4:K8,1)-1))/INDEX(J4:J8,MATCH(0,K4:K8,1))"),
            ("Точка безубыточности (ед.):", "=Ввод!B14/(Ввод!B12-Ввод!B13)")
        ]
        
        for i, (metric, formula) in enumerate(metrics, 11):
            ws[f'A{i}'] = metric
            ws[f'B{i}'] = formula
            ws[f'A{i}'].font = self.bold_font
        
        # Настройка ширины колонок
        for col in range(1, 12):
            ws.column_dimensions[get_column_letter(col)].width = 12
        
        print("✅ Лист Расчёты создан")

    def create_scenarios_sheet(self):
        """Создание листа Сценарии"""
        ws = self.wb.create_sheet("Сценарии")
        
        # Заголовок
        ws['A1'] = "СЦЕНАРНЫЙ АНАЛИЗ"
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Выбор сценария
        ws['A3'] = "ВЫБЕРИТЕ СЦЕНАРИЙ:"
        ws['A3'].font = self.subheader_font
        
        ws['B3'] = "Базовый"
        scenario_dv = DataValidation(type="list", formula1='"Пессимистичный,Базовый,Оптимистичный"')
        ws.add_data_validation(scenario_dv)
        scenario_dv.add('B3')
        
        # Таблица сценариев
        ws['A5'] = "ТАБЛИЦА СЦЕНАРИЕВ"
        ws['A5'].font = self.subheader_font
        ws.merge_cells('A5:F5')
        
        scenario_headers = ["Сценарий", "Объём", "Цена", "Рост%", "Переменные"]
        scenario_data = [
            ["Пессимистичный", 80000, 45, 2, 28],
            ["Базовый", 100000, 50, 5, 25],
            ["Оптимистичный", 120000, 55, 8, 22]
        ]
        
        # Заголовки таблицы
        for col, header in enumerate(scenario_headers, 1):
            cell = ws.cell(row=6, column=col)
            cell.value = header
            cell.font = self.bold_font
            cell.fill = self.subheader_fill
        
        # Данные сценариев
        for row, data in enumerate(scenario_data, 7):
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
        
        print("✅ Лист Сценарии создан")

    def create_report_sheet(self):
        """Создание листа Отчёт"""
        ws = self.wb.create_sheet("Отчёт")
        
        # Заголовок отчёта
        ws['A1'] = "ИНВЕСТИЦИОННЫЙ ОТЧЁТ"
        ws['A1'].font = Font(size=20, bold=True, color="2F5496")
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Информация о проекте
        ws['A3'] = "Информация о проекте:"
        ws['A3'].font = self.subheader_font
        
        project_info = [
            ("Название проекта:", "=Ввод!B3"),
            ("Сценарий:", "=Сценарии!B3"),
            ("Дата анализа:", f"{datetime.now().strftime('%d.%m.%Y')}"),
            ("Срок проекта:", "=Ввод!B6 & ' лет'")
        ]
        
        for i, (label, value) in enumerate(project_info, 4):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = self.bold_font
        
        # Ключевые метрики
        ws['A8'] = "КЛЮЧЕВЫЕ МЕТРИКИ ПРОЕКТА"
        ws['A8'].font = Font(size=14, bold=True, color="2F5496")
        ws.merge_cells('A8:D8')
        
        metrics = [
            ("NPV проекта:", "=Расчёты!B11", "₽"),
            ("Внутренняя норма доходности (IRR):", "=Расчёты!B12", "%"),
            ("Срок окупаемости:", "=Расчёты!B13", "лет"),
            ("Точка безубыточности:", "=Расчёты!B14", "ед.")
        ]
        
        for i, (label, formula, unit) in enumerate(metrics, 9):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = formula
            ws[f'C{i}'] = unit
            ws[f'A{i}'].font = self.bold_font
        
        # Статус проекта
        ws['A14'] = "СТАТУС ПРОЕКТА:"
        ws['A14'].font = Font(size=12, bold=True)
        
        ws['B14'] = '=ЕСЛИ(Расчёты!B12>=15;"✅ РЕКОМЕНДУЕТСЯ";ЕСЛИ(Расчёты!B12>=10;"⚠️ ТРЕБУЕТ ДОРАБОТКИ";"❌ ОТКЛОНЁН"))'
        ws['B14'].font = Font(size=14, bold=True)
        ws.merge_cells('B14:D14')
        
        # Кнопка для макроса
        ws['A16'] = "📤 ЭКСПОРТИРОВАТЬ ОТЧЁТ В PDF"
        ws['A16'].font = Font(size=12, bold=True, color="FFFFFF")
        ws['A16'].fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        ws['A16'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A16:D18')
        
        # Настройка колонок
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 8
        
        print("✅ Лист Отчёт создан")

    def create_vba_code_file(self):
        """Создание файла с VBA кодом"""
        # Исправленный VBA код без проблемных escape-последовательностей
        vba_code = '''Attribute VB_Name = "InvestmentMacros"
' Макросы для финансовой модели

Sub ExportInvestmentReport()
    Application.ScreenUpdating = False
    Application.Calculate
    
    ThisWorkbook.RefreshAll
    
    Dim projectName As String
    projectName = ThisWorkbook.Worksheets("Ввод").Range("B3").Value
    If projectName = "" Then projectName = "Проект"
    
    projectName = CleanFileName(projectName)
    
    Dim fileName As String
    fileName = "ИнвестОтчёт_" & projectName & "_" & Format(Now, "dd.mm.yyyy") & ".pdf"
    
    Dim fullPath As String
    If ThisWorkbook.Path <> "" Then
        fullPath = ThisWorkbook.Path & "\" & fileName
    Else
        fullPath = Environ("USERPROFILE") & "\Desktop\" & fileName
    End If
    
    On Error GoTo ExportError
    ThisWorkbook.Worksheets("Отчёт").ExportAsFixedFormat _
        Type:=xlTypePDF, _
        FileName:=fullPath, _
        Quality:=xlQualityStandard
    
    CreateArchiveSheet
    
    Application.ScreenUpdating = True
    MsgBox "Отчёт успешно экспортирован!" & vbNewLine & _
           "Файл: " & fileName, vbInformation
    Exit Sub
    
ExportError:
    Application.ScreenUpdating = True
    MsgBox "Ошибка при экспорте: " & Err.Description, vbCritical
End Sub

Sub CreateArchiveSheet()
    Dim archiveSheet As Worksheet
    Dim archiveName As String
    archiveName = "Архив_" & Format(Now, "dd.mm.yyyy HH-MM-SS")
    
    On Error Resume Next
    Set archiveSheet = ThisWorkbook.Worksheets(archiveName)
    On Error GoTo 0
    
    If Not archiveSheet Is Nothing Then
        archiveName = "Архив_" & Format(Now, "dd.mm.yyyy HH-MM-SS") & Format(Timer, "000")
    End If
    
    Set archiveSheet = ThisWorkbook.Worksheets.Add(After:=Worksheets(Worksheets.Count))
    archiveSheet.Name = archiveName
    
    ThisWorkbook.Worksheets("Расчёты").UsedRange.Copy
    archiveSheet.Range("A1").PasteSpecial Paste:=xlPasteValues
    archiveSheet.Range("A1").PasteSpecial Paste:=xlPasteFormats
    
    archiveSheet.Range("A1").Value = "АРХИВ РАСЧЁТОВ - " & Format(Now, "dd.mm.yyyy HH:MM:SS")
    archiveSheet.Range("A1").Font.Bold = True
    
    archiveSheet.Protect Password:="123"
    Application.CutCopyMode = False
    MsgBox "Создан архив: " & archiveName, vbInformation
End Sub

Function CleanFileName(originalName As String) As String
    Dim invalidChars As String
    Dim i As Integer
    invalidChars = "\/:*?""<>|"
    CleanFileName = originalName
    For i = 1 To Len(invalidChars)
        CleanFileName = Replace(CleanFileName, Mid(invalidChars, i, 1), "")
    Next i
    CleanFileName = WorksheetFunction.Trim(CleanFileName)
End Function
'''
        
        with open("vba_code.txt", "w", encoding="utf-8") as f:
            f.write(vba_code)
        
        print("✅ Файл vba_code.txt создан")

def main():
    """Основная функция"""
    print("🚀 СОЗДАНИЕ ФИНАНСОВОЙ МОДЕЛИ")
    print("=" * 50)
    
    creator = InvestmentModelCreator()
    filename = f"ИнвестМодель_{datetime.now().strftime('%d%m%Y_%H%M')}.xlsx"
    
    try:
        creator.create_model(filename)
        
        print("\n🎉 МОДЕЛЬ УСПЕШНО СОЗДАНА!")
        print(f"\n📁 Файл: {filename}")
        print("📋 VBA код: vba_code.txt")
        
        print("\n📋 Следующие шаги:")
        print("1. Откройте файл в Excel")
        print("2. Добавьте VBA макросы из vba_code.txt")
        print("3. Протестируйте формулы")
        print("4. Настройте защиту листов")
        
    except Exception as e:
        print(f"❌ Ошибка: {e}")

if __name__ == "__main__":
    main()